# =======================
# Report Service — CSV, Excel, PDF generation
# =======================
"""
Generates in-memory reports from platform data.
Supports CSV, Excel (xlsx), and PDF formats.
"""

import csv
import io
from datetime import datetime
from typing import Dict, List


class ReportService:
    def __init__(self, db_pool):
        self.pool = db_pool

    # ---------- Data fetchers ----------

    async def _fetch_orders(self) -> List[Dict]:
        if not self.pool:
            return []
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT o.order_number, c.name AS customer_name, o.status,
                       o.order_date, o.total_amount, o.payment_terms,
                       o.shipping_method
                FROM orders o
                LEFT JOIN customers c ON c.id = o.customer_id
                ORDER BY o.order_date DESC
                LIMIT 1000
            """)
            return [dict(r) for r in rows]

    async def _fetch_inventory(self) -> List[Dict]:
        if not self.pool:
            return []
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT p.sku, p.name, p.manufacturer, p.category,
                       i.warehouse_code, i.quantity_on_hand, i.quantity_reserved,
                       i.reorder_point, i.bin_location
                FROM inventory i
                JOIN products p ON p.id = i.product_id
                ORDER BY p.sku
                LIMIT 5000
            """)
            return [dict(r) for r in rows]

    async def _fetch_invoices(self) -> List[Dict]:
        if not self.pool:
            return []
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT inv.invoice_number, c.name AS customer_name, inv.status,
                       inv.invoice_date, inv.due_date, inv.total_amount,
                       inv.amount_paid, inv.balance_due
                FROM invoices inv
                LEFT JOIN customers c ON c.id = inv.customer_id
                ORDER BY inv.invoice_date DESC
                LIMIT 1000
            """)
            return [dict(r) for r in rows]

    async def _fetch_sales(self, period: str = "month") -> List[Dict]:
        trunc_map = {
            "day": "day",
            "week": "week",
            "month": "month",
        }
        trunc = trunc_map.get(period, "month")
        if not self.pool:
            return []
        async with self.pool.acquire() as conn:
            rows = await conn.fetch(f"""
                SELECT date_trunc('{trunc}', o.order_date) AS period,
                       COUNT(*) AS order_count,
                       SUM(o.total_amount) AS revenue
                FROM orders o
                WHERE o.status NOT IN ('cancelled', 'draft')
                GROUP BY period
                ORDER BY period DESC
                LIMIT 365
            """)
            return [dict(r) for r in rows]

    # ---------- CSV ----------

    def generate_csv(self, data: List[Dict]) -> bytes:
        if not data:
            return b""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        for row in data:
            writer.writerow({k: self._format_value(v) for k, v in row.items()})
        return output.getvalue().encode("utf-8")

    # ---------- Excel ----------

    def generate_xlsx(self, data: List[Dict], sheet_name: str = "Report") -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb_bytes = io.BytesIO()
            wb.save(wb_bytes)
            return wb_bytes.getvalue()

        # Header row
        headers = list(data[0].keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=self._format_value(row[key]))

        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ---------- PDF ----------

    def generate_pdf(self, data: List[Dict], title: str = "Report") -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter),
                                leftMargin=0.5 * inch, rightMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"]
        ))
        elements.append(Spacer(1, 12))

        if not data:
            elements.append(Paragraph("No data available.", styles["Normal"]))
            doc.build(elements)
            return output.getvalue()

        # Table
        headers = list(data[0].keys())
        header_labels = [h.replace("_", " ").title() for h in headers]
        table_data = [header_labels]

        for row in data[:500]:  # Limit rows for PDF
            table_data.append([
                str(self._format_value(row.get(k, "")))[:50]
                for k in headers
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        elements.append(table)
        doc.build(elements)
        return output.getvalue()

    # ---------- Helpers ----------

    @staticmethod
    def _format_value(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M")
        return value
