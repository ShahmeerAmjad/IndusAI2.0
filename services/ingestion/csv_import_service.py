"""CSV/Excel product import service.

Parses uploaded files, maps columns to product schema,
and feeds products into the same pipeline as web scraping.
"""

import csv
import io
import logging

logger = logging.getLogger(__name__)

COLUMN_ALIASES = {
    "product_name": "name",
    "product name": "name",
    "supplier": "manufacturer",
    "mfg": "manufacturer",
    "cas": "cas_number",
    "cas #": "cas_number",
    "cas_no": "cas_number",
    "desc": "description",
    "industries": "industry",
}


class CSVImportService:
    def __init__(self, pipeline, llm_router=None):
        self._pipeline = pipeline
        self._llm = llm_router

    async def parse_file(self, file_obj: io.BytesIO, filename: str) -> list[dict]:
        """Parse CSV or Excel file into list of product dicts."""
        if filename.endswith((".xlsx", ".xls")):
            return self._parse_excel(file_obj)
        return self._parse_csv(file_obj)

    def _parse_csv(self, file_obj: io.BytesIO) -> list[dict]:
        text = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        column_map = {col: COLUMN_ALIASES.get(col.lower().strip(), col.lower().strip())
                      for col in (reader.fieldnames or [])}
        products = []
        for row in reader:
            product = {}
            for orig_col, mapped_col in column_map.items():
                val = row.get(orig_col, "").strip()
                if val:
                    if mapped_col == "industry":
                        product["industries"] = [i.strip() for i in val.split(",")]
                    else:
                        product[mapped_col] = val
            if product.get("name"):
                products.append(product)
        return products

    def _parse_excel(self, file_obj: io.BytesIO) -> list[dict]:
        from openpyxl import load_workbook
        wb = load_workbook(file_obj, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        column_map = {h: COLUMN_ALIASES.get(h, h) for h in headers}
        products = []
        for row in rows[1:]:
            product = {}
            for i, val in enumerate(row):
                if i < len(headers) and val is not None:
                    mapped = column_map.get(headers[i], headers[i])
                    val_str = str(val).strip()
                    if mapped == "industry":
                        product["industries"] = [v.strip() for v in val_str.split(",")]
                    else:
                        product[mapped] = val_str
            if product.get("name"):
                products.append(product)
        wb.close()
        return products

    async def dry_run(self, file_obj: io.BytesIO, filename: str) -> dict:
        """Preview import: return first 5 rows and column mapping."""
        products = await self.parse_file(file_obj, filename)
        columns = list(products[0].keys()) if products else []
        return {
            "columns": columns,
            "sample_rows": products[:5],
            "total_rows": len(products),
        }

    async def import_products(self, products: list[dict],
                               on_progress=None) -> dict:
        """Import parsed products through the seed pipeline."""
        _emit = on_progress or (lambda e: None)
        stats = {"products_created": 0, "products_updated": 0,
                 "tds_stored": 0, "sds_stored": 0,
                 "industries_linked": 0, "errors": 0}

        for i, product in enumerate(products):
            _emit({"stage": "importing", "current": i + 1,
                   "total": len(products), "product": product.get("name", "")})
            try:
                await self._pipeline._process_product(product, stats, _emit)
            except Exception as e:
                logger.warning("Import failed for %s: %s", product.get("name"), e)
                stats["errors"] += 1

        return stats
